from tkinter import *
import openpyxl
import os

FONT = ("Arial", "12")


def calculate_commission():
    qty = float(quantity.get())
    percentage = float(commission_per.get())
    gross = float(gross_sales.get())
    labor_rate = float(labor_charge.get())
    commission = percentage/100 * gross
    net_amount = gross - (commission + qty * labor_rate)
    commission_amount.config(text=commission)
    net_sales.config(text=net_amount)
    save_to_exel(qty, percentage, gross, labor_rate, commission, net_amount)

def save_to_exel(qty, pct, gross, labor, com, net):

    file_path = 'data.xlsx'

    # Check if the file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    else:
        # create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # add data to the notebook
        ws['A1'] = 'Quantity'
        ws['B1'] = 'Percentage'
        ws['C1'] = 'Gross'
        ws['D1'] = 'Labor Rate'
        ws['E1'] = 'Commission'
        ws['F1'] = 'Net Amount'

    # Add data to the next empty row
    ws.append([qty, pct, gross, labor, com, net])

    # Save the workbook
    wb.save(file_path)


def reset():
    quantity.delete(0, END)
    commission_per.delete(0, END)
    gross_sales.delete(0, END)
    labor_charge.delete(0, END)


window = Tk()
window.title("Commission Calculator")
window.config(padx=30, pady=30, bg="gray")

# Entry for Quantity
quantity = Entry()
quantity.grid(column=1, row=0)

# Label for unit
unit = Label(text="Units", bg="gray", font=FONT)
unit.grid(column=2, row=0)

# Entry for gross amount
gross_sales = Entry()
gross_sales.grid(column=1, row=1)

# Label for gross sales
gross_label = Label(text="Gross Amount", bg="gray", font=FONT)
gross_label.grid(column=2, row=1)

# Commission % Entry
commission_per = Entry()
commission_per.grid(column=1, row=2)

# Commission % label
label_per = Label(text="Commission %", bg="gray", font=FONT)
label_per.grid(column=2, row=2)

# Labor charge per unit
labor_charge = Entry()
labor_charge.grid(column=1, row=3)

# Label for labor charge
labor_label = Label(text="Per Unit Labor Charge", bg="gray", font=FONT)
labor_label.grid(column=2, row=3)

# Commission Calculated
commission_amount = Label(text=0, bg="gray", fg="yellow", font=FONT)
commission_amount.grid(column=1, row=4)

# Commission Label
commission_label = Label(text="Commission", bg="gray", fg="red", font=FONT)
commission_label.grid(column=2, row=4)

# Label for net amount
net_sales = Label(text=0, bg="gray", fg="yellow", font=FONT)
net_sales.grid(column=1, row=5)

# Label for net sales
net_label = Label(text="Net Amount", bg="gray", fg="red", font=FONT)
net_label.grid(column=2, row=5)

# Calculate button
calculate_button = Button(text="Calculate", command=calculate_commission)
calculate_button.grid(column=1, row=6)
calculate_button.config(padx=10, pady=10)

# Reset button
reset_button = Button(text="Reset", command=reset)
reset_button.grid(column=2, row=6)
reset_button.config(padx=10, pady=10)

window.mainloop()
